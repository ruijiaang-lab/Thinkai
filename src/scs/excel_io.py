"""Excel 读取、表头识别、字段映射（确定性启发式）。

说明：表头识别是基于关键词规则的确定性程序，不是模型推断；
识别结果只是"建议"，最终映射必须由用户确认后保存。
"""
from __future__ import annotations

import unicodedata
from datetime import date, datetime, time
from pathlib import Path
from typing import Optional

import openpyxl

# 字段 -> 表头关键词（按优先级排列，精确等价优先于包含匹配）
FIELDS = ["uid", "amount", "order_no", "date", "name", "status"]
FIELD_KEYWORDS = {
    "uid": ["uid", "用户id", "用户 id", "工会id", "主播id", "账号", "memberid",
            "member", "编号", "id"],
    "amount": ["结算金额", "应结金额", "金额", "佣金", "分成", "结算", "amount"],
    "order_no": ["订单号", "订单编号", "单号", "订单", "order"],
    "date": ["日期", "结算日期", "时间", "周期", "date"],
    "name": ["昵称", "姓名", "主播", "name"],
    "status": ["状态", "status"],
}
FIELD_LABELS = {
    "uid": "UID", "amount": "金额", "order_no": "订单号",
    "date": "日期", "name": "昵称", "status": "状态",
}


def cell_to_text(v) -> str:
    """单元格值 -> 展示文本（日期/数字做友好格式化）。"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        if v.time() == time(0, 0):
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _header_text(v) -> str:
    if v is None:
        return ""
    return unicodedata.normalize("NFKC", str(v)).strip().lower()


def read_rows(path, sheet: Optional[str] = None) -> list:
    """读取第一个（或指定）工作表，返回二维数组。仅支持 .xlsx。"""
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"暂只支持 .xlsx 格式（收到：{path.name}）。"
                         f"如为旧版 .xls，请先用 Excel/WPS 另存为 .xlsx。")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    return rows


def detect_header_row(rows: list, max_scan: int = 15) -> Optional[int]:
    """在前 max_scan 行里找命中字段关键词最多的行作为表头。

    至少命中 2 类字段才认可；找不到返回 None（交人工指定）。
    """
    best_idx, best_score = None, 0
    for i, row in enumerate(rows[:max_scan]):
        cells = [_header_text(c) for c in row if c is not None and str(c).strip()]
        if not cells:
            continue
        score = 0
        for kws in FIELD_KEYWORDS.values():
            if any(any(kw in c for kw in kws) for c in cells):
                score += 1
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx if best_score >= 2 else None


def guess_mapping(rows: list, header_idx: int) -> dict:
    """对表头行做字段匹配，返回建议映射 {field: {column, header}}。"""
    header = rows[header_idx]
    cells = [(i, _header_text(c), c) for i, c in enumerate(header)
             if c is not None and str(c).strip()]
    mapping: dict = {}
    for field, kws in FIELD_KEYWORDS.items():
        # 精确等价优先
        hit = next(((i, raw) for i, norm, raw in cells if norm in kws), None)
        # 其次包含匹配
        if hit is None:
            hit = next(((i, raw) for i, norm, raw in cells
                        if any(kw in norm for kw in kws)), None)
        if hit:
            mapping[field] = {"column": hit[0], "header": str(hit[1]).strip()}
    return mapping


def apply_mapping(rows: list, header_idx: int, mapping: dict) -> list:
    """按映射抽取数据行，返回 [{"row_no": Excel行号, "uid": 原始值, ...}]。

    mapping 支持两种写法：
      {"uid": {"column": 0, "header": "UID"}}（save-mapping 的格式）
      {"uid": "UID"}（手写简写）
    优先按表头名匹配列，匹配不到再用列号兜底。
    """
    header = rows[header_idx]

    def col_for(field: str) -> Optional[int]:
        m = mapping.get(field)
        if m is None:
            return None
        target = ""
        col = None
        if isinstance(m, str):
            target = m
        elif isinstance(m, dict):
            target = str(m.get("header") or "")
            c = m.get("column")
            col = c if isinstance(c, int) else None
        if target:
            target_norm = unicodedata.normalize("NFKC", target).strip().lower()
            for i, c in enumerate(header):
                if c is not None and _header_text(c) == target_norm:
                    return i
        return col

    cols = {f: col_for(f) for f in FIELDS}
    records = []
    for offset, row in enumerate(rows[header_idx + 1:]):
        row_no = header_idx + 2 + offset  # 1 基 Excel 行号
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        rec = {"row_no": row_no}
        for f, ci in cols.items():
            rec[f] = row[ci] if (ci is not None and ci < len(row)) else None
        records.append(rec)
    return records


def mapping_summary(mapping: dict) -> str:
    """人话版映射说明。"""
    lines = []
    for f in FIELDS:
        m = mapping.get(f)
        if not m:
            continue
        header = m.get("header") if isinstance(m, dict) else m
        col = m.get("column") if isinstance(m, dict) else None
        col_txt = f"第 {col + 1} 列" if isinstance(col, int) else ""
        lines.append(f"  {FIELD_LABELS[f]}（{f}）← 表头「{header}」{col_txt}")
    missing = [FIELD_LABELS[f] for f in ("uid", "amount") if f not in mapping]
    txt = "\n".join(lines) or "  （无）"
    if missing:
        txt += f"\n  ⚠ 未识别到必需字段：{'、'.join(missing)}"
    return txt
