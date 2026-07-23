"""结算预览 Excel 生成（给财务人工检查用）。"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .engine import BUCKET_LABELS, BUCKET_ORDER, TO_SETTLE

_COLUMNS = ["行号", "UID", "订单号", "金额", "台账参考", "日期", "昵称", "原因"]
_WIDTHS = [8, 16, 16, 12, 12, 12, 12, 46]


def _write_bucket_sheet(wb, title: str, items):
    ws = wb.create_sheet(title)
    ws.append(_COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for it in items:
        amount = None
        if it.amount not in (None, ""):
            try:
                amount = float(it.amount)
            except ValueError:
                amount = it.amount
        ref = None
        if it.ref_amount not in (None, ""):
            try:
                ref = float(it.ref_amount)
            except ValueError:
                ref = it.ref_amount
        ws.append([it.row_no, it.uid or it.raw_uid, it.order_no or "",
                   amount, ref, it.date, it.name, it.reason])
        if isinstance(amount, float):
            ws.cell(row=ws.max_row, column=4).number_format = "0.00"
        if isinstance(ref, float):
            ws.cell(row=ws.max_row, column=5).number_format = "0.00"
    for i, w in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_summary_sheet(wb, result):
    ws = wb.create_sheet("汇总", 0)
    s = result.summary
    rows = [
        ("源文件", Path(result.source_file).name),
        ("批次号", result.batch_id),
        ("生成时间", result.generated_at),
        ("原始数据行数", s.get("total_rows", 0)),
        ("待结算笔数", s.get(TO_SETTLE, 0)),
        ("待结算金额", s.get("to_settle_amount", "0.00")),
        ("已结算跳过", s.get("already_settled", 0)),
        ("批内重复", s.get("duplicate_in_batch", 0)),
        ("缺UID", s.get("missing_uid", 0)),
        ("金额不一致", s.get("amount_mismatch", 0)),
        ("需人工复核", s.get("manual_review", 0)),
    ]
    ws.append(["指标", "数值"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for k, v in rows:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40


def write_preview_excel(result, path) -> Path:
    """按分桶生成多 Sheet 预览表。Sheet 名不超过 31 字符（Excel 限制）。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_summary_sheet(wb, result)
    for bucket in BUCKET_ORDER:
        _write_bucket_sheet(wb, BUCKET_LABELS[bucket], result.bucket_items(bucket))
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
